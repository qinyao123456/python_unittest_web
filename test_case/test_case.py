#文件  测试用例 登录和选择商品 结合在一起  知识点 unittest/pytest  单元测试框架 用例管理框架
import unittest
from selenium import webdriver
from page.login import LoginPage
from page.choice_page import ChoicePage
from ddt import ddt,data,unpack,file_data
import yaml
import openpyxl
import os
import HTMLTestRunner
from common.excel_util import ExcelUtil
#file_data是读取yml文件的装饰器
def getData():
    list = []
    with open('data.txt','r',encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            list.append(line.strip('\n').split(','))
        return list
#读取excel
#openpyxl
def read_excel():
    workbook = openpyxl.load_workbook(os.path.abspath(os.path.dirname(__file__))+'/data.xlsx')
    sheet = workbook['login']
    print(sheet.max_row,sheet.max_column)
    allList = []
    for row in range(2,sheet.max_row+1):           #行数
        tempList = []
        for col in range(1,sheet.max_column+1):
            print(row,col)
            tempList.append(sheet.cell(row,col).value)
        allList.append(tempList)
    print(allList)
    return allList

#必须要继承unittest.TestCase
#ddt(data-driver-test)数据驱动，针对单侧测试有多组数据
@ddt   #在class前定义@ddt 表明在这个类之中会用这个ddt(装饰器)
class TestCase(unittest.TestCase):
    #用例执行前要做的事情 写 test_ 一个test方法 代表一条用例 登录  找商品  用例执行前干什么  打开浏览器
    #测试用例 def test_()  必须以test开头
    #test suite 测试套件，也称之为测试用例集
    #test runner 运行器，一般通过runnner去执行suite集合
    #unittest 运行机制：通过在main 函数中，调用unittest.main()运行所有的测试用例
    @classmethod
    def setUpClass(cls) -> None:
        cls.driver = webdriver.Chrome()
        cls.lp =LoginPage(cls.driver)
        cls.cp =ChoicePage(cls.driver)



    #用例执行后要做的事
    @classmethod
    def tearDownClass(cls) -> None:
        pass

#用例卸载excel 自动化测试 回归测试 有新功能的时候，都要先跑一次回归测试
    @data('你好','123456')
    def test_01(self,txt):
        print(txt)
    @data((10,20),(30,40))    #用于设定参数
    @unpack   #用于解析参数
    def test_02(self,txt,arm):
        print(txt)
        print(arm)
    @data(*getData())
    @unpack
    def test_03(self,txt,arm):
        print(txt)
        print(arm)
    @data(read_excel())
    def test_04(self,txt):
        print(txt)            #执行一次 [[1, 'admin', 123456], [2, 'admin', 1234567], [3, 'admin', 12345678]]
    @data(*read_excel())
    def test_05(self,txt):
        print(txt)               #执行三次[1, 'admin', 123456] [2, 'admin', 1234567] [3, 'admin', 12345678]
    @data(*read_excel())
    @unpack
    def test_06(self,txt1,txt2,txt3):
        print(txt1,txt2,txt3)
    #从common/excel_util获取数据
    @data(*ExcelUtil().read_excel())
    def test_07(self,txt):
        print(txt)
if __name__ == '__main__':
    # unittest.main()
    testcases= unittest.defaultTestLoader.discover(os.path.abspath(os.path.dirname(__file__)),"*.py")
    fileName= open(os.path.abspath(os.path.dirname(__file__))+"/report.html","wb")
    runner= HTMLTestRunner.HTMLTestRunner(stream=fileName,title = "python练习")
    runner.run(testcases)