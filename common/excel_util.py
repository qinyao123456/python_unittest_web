import openpyxl
import os
class ExcelUtil:
    #./ 当前目录
    # openpyxl.load_workbook('./data/data.xlsx')
    #获得项目路径

    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def read_excel(self):
        #openpyxl,xlrd
        wb = openpyxl.load_workbook(self.get_object_path()+'data/data.xlsx')
        #获得sheet对象
        sheet = wb['login']
        allList = []
        #获得excel的行数和列数
        for row in range(2,sheet.max_row+1):
            tempList = []
            for col in range(1,sheet.max_column+1):
                tempList.append(sheet.cell(row,col).value)
            allList.append(tempList)
        return allList




if __name__ == '__main__':
    print(os.getcwd())
    ExcelUtil().get_object_path()
    print(ExcelUtil().read_excel())
